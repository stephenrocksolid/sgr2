import csv
import io
import json
import re
import time
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional, Tuple, Any
from django.core.files.base import ContentFile
from django.db import transaction
from django.contrib.auth.models import User
from django.db.models import Q
from openpyxl import load_workbook
from .models import ImportBatch, ImportLog, SavedImportMapping, ImportRow
from inventory.models import Machine, Engine, Part, SGEngine, Vendor, SGVendor, PartVendor, MachineEngine, EnginePart, MachinePart, PartAttribute, PartAttributeValue, PartAttributeChoice

def get_valid_model_fields(model_class):
    """Get a set of valid field names for a Django model."""
    return {field.name for field in model_class._meta.get_fields()}

def filter_valid_fields(data, model_class):
    """Filter data to only include fields that exist on the model."""
    valid_fields = get_valid_model_fields(model_class)
    filtered_data = {k: v for k, v in data.items() if k in valid_fields}
    
    # Log filtered out fields for debugging
    filtered_out = {k: v for k, v in data.items() if k not in valid_fields}
    if filtered_out:
        print(f"Warning: Filtered out invalid fields for {model_class.__name__}: {list(filtered_out.keys())}")
    
    return filtered_data

def normalize(s):
    """Normalize string for comparison."""
    return (s or '').strip()

def ci_get_or_create_vendor(name, vendor_data=None):
    """Get or create vendor by case-insensitive name."""
    n = normalize(name)
    if not n:
        # Return Unknown default vendor when name is None/empty
        unknown_vendor = Vendor.objects.get_or_create(
            name='Unknown',
            defaults={
                'notes': 'Default record for imports without vendor information'
            }
        )[0]
        return unknown_vendor
    
    v = Vendor.objects.filter(name__iexact=n).first()
    if v:
        return v
    
    # Create new vendor with additional data if provided
    vendor_defaults = {'name': n}
    if vendor_data:
        vendor_defaults.update({
            'website': vendor_data.get('vendor_website', ''),
            'contact_name': vendor_data.get('vendor_contact_name', ''),
            'email': vendor_data.get('vendor_contact_email', ''),
            'phone': vendor_data.get('vendor_contact_phone', ''),
            'notes': vendor_data.get('vendor_notes', ''),
        })
    
    return Vendor.objects.create(**vendor_defaults)

def _nz(s):
    """Normalize string for comparison."""
    return (s or "").strip().lower()

def build_engine_key(row, mapping):
    """
    Returns (make, model, identifier) normalized from the row using Step-3 mapping.
    Supports 'identifier' as the mapped key.
    """
    mk_key = mapping.get('engine_make')    # mapped header name
    md_key = mapping.get('engine_model')
    id_key = mapping.get('identifier')
    make = _nz(row.get(mk_key))
    model = _nz(row.get(md_key))
    ident = _nz(row.get(id_key))
    return (make, model, ident)

def is_valid_engine_key(tup):
    """Require all three pieces to be non-empty."""
    return all(bool(x) for x in tup)

def validate_and_truncate_fields(data, model_class, batch=None, row_number=None):
    """Validate and truncate field values to match model constraints."""
    if not data:
        return data
    
    validated_data = {}
    
    # Get model field information
    model_fields = {field.name: field for field in model_class._meta.get_fields()}
    
    for field_name, value in data.items():
        if field_name not in model_fields:
            validated_data[field_name] = value
            continue
            
        field = model_fields[field_name]
        
        # Skip if value is None or empty
        if value is None or value == '':
            validated_data[field_name] = value
            continue
            
        # Convert to string for validation
        str_value = str(value)
        
        # Check for CharField with max_length
        if hasattr(field, 'max_length') and field.max_length:
            if len(str_value) > field.max_length:
                truncated_value = str_value[:field.max_length]
                validated_data[field_name] = truncated_value
                
                # Log truncation warning
                warning_msg = f"Truncated {field_name} from {len(str_value)} to {field.max_length} characters: '{str_value}' -> '{truncated_value}'"
                print(f"Warning: {warning_msg}")
                
                # Log to database if batch info is available
                if batch and row_number is not None:
                    ImportLog.objects.create(
                        batch=batch,
                        level='warning',
                        message=warning_msg,
                        row_number=row_number
                    )
            else:
                validated_data[field_name] = value
        else:
            validated_data[field_name] = value
    
    return validated_data

# Import Celery task decorator with fallback
try:
    from celery import shared_task
    CELERY_AVAILABLE = True
except ImportError:
    # Fallback decorator when Celery is not available
    def shared_task(*args, **kwargs):
        def decorator(func):
            return func
        return decorator
    CELERY_AVAILABLE = False

@shared_task
def test_task():
    return "Celery is working!"

def _process_import_batch_internal(batch_id, task_instance=None):
    """Internal function to process an import batch. Can be called directly or from Celery task."""
    start_time = time.time()
    
    try:
        batch = ImportBatch.objects.get(id=batch_id)
        batch.status = 'processing'
        batch.progress_percentage = 0
        batch.save()
        
        # Update Celery task state
        if CELERY_AVAILABLE and task_instance and hasattr(task_instance, 'update_state'):
            task_instance.update_state(
                state='PROGRESS',
                meta={'rows': 0, 'created': 0, 'errors': 0, 'status': 'Starting import...'}
            )
        
        # Log start
        ImportLog.objects.create(
            batch=batch,
            level='info',
            message=f"Starting import processing for {batch.original_filename}"
        )
        
        # Get mapping configuration
        if not batch.mapping:
            raise Exception("No mapping configuration found for this batch.")
        
        mapping = batch.mapping
        
        # Process file based on type
        if batch.is_csv():
            success_count, error_count = process_csv_import(batch, mapping, task_instance)
        elif batch.is_xlsx():
            success_count, error_count = process_xlsx_import(batch, mapping, task_instance)
        else:
            raise Exception(f"Unsupported file type: {batch.file_type}")
        
        # Refresh batch to check if it was cancelled during processing
        batch.refresh_from_db()
        
        # Update batch status (only if not already cancelled)
        if batch.status != 'cancelled':
            batch.status = 'completed'
            batch.progress_percentage = 100
            batch.processed_rows = success_count + error_count
            batch.save()
            
            # Final Celery task state update
            if CELERY_AVAILABLE and task_instance and hasattr(task_instance, 'update_state'):
                task_instance.update_state(
                    state='SUCCESS',
                    meta={
                        'rows': success_count + error_count,
                        'created': success_count,
                        'errors': error_count,
                        'status': 'Import completed successfully'
                    }
                )
            
            # Log completion
            ImportLog.objects.create(
                batch=batch,
                level='info',
                message=f"Import completed in {time.time() - start_time:.2f}s. {success_count} records processed successfully, {error_count} errors."
            )
        else:
            # Import was cancelled
            batch.processed_rows = success_count + error_count
            batch.save()
            return success_count, error_count
        
        return {
            'success': True,
            'processed_rows': success_count + error_count,
            'success_count': success_count,
            'error_count': error_count,
            'processing_time': time.time() - start_time
        }
        
    except Exception as e:
        # Update batch status to failed
        try:
            batch = ImportBatch.objects.get(id=batch_id)
            batch.status = 'failed'
            batch.error_message = str(e)
            batch.save()
            
            # Update Celery task state
            if CELERY_AVAILABLE and task_instance and hasattr(task_instance, 'update_state'):
                task_instance.update_state(
                    state='FAILURE',
                    meta={'status': f'Import failed: {str(e)}'}
                )
            
            ImportLog.objects.create(
                batch=batch,
                level='error',
                message=f"Import failed: {str(e)}"
            )
        except:
            pass
        
        return {
            'success': False,
            'error': str(e)
        }

@shared_task(bind=True, time_limit=1800, soft_time_limit=1500)
def process_import_batch(self, batch_id):
    """Celery task wrapper for processing an import batch."""
    return _process_import_batch_internal(batch_id, task_instance=self)

def process_import_batch_sync(batch_id):
    """Synchronous wrapper for processing an import batch (for threading/testing)."""
    return _process_import_batch_internal(batch_id, task_instance=None)

def process_csv_import(batch, mapping, task_instance=None):
    """Process CSV import in chunks with streaming."""
    success_count = 0
    error_count = 0
    
    # Preload existing engine keys for fast duplicate detection
    existing_engine_keys = set()
    for make, model, ident in Engine.objects.values_list('engine_make', 'engine_model', 'identifier'):
        existing_engine_keys.add((_nz(make), _nz(model), _nz(ident)))
    
    # Track keys we create/update in this run to avoid dupes within the batch
    batch_engine_keys = set()
    
    # Use streaming approach to avoid loading entire file into memory
    chunk_size = mapping.chunk_size or 1000  # Default to 1000 if not set
    
    # Read file in streaming fashion
    with batch.file.open('rb') as f:
        # Detect encoding and delimiter if needed
        file_content_sample = f.read(8192)  # Read first 8KB for detection
        f.seek(0)  # Reset to beginning
        
        # Process CSV with streaming
        import csv
        import io
        
        # Decode the file content
        text_content = file_content_sample.decode(batch.encoding or 'utf-8', errors='ignore')
        
        # Get headers from first line
        first_line = text_content.split('\n')[0]
        headers = [h.strip() for h in first_line.split(batch.delimiter or ',')]
        
        # Stream process the file
        f.seek(0)
        text_stream = io.TextIOWrapper(f, encoding=batch.encoding or 'utf-8')
        csv_reader = csv.reader(text_stream, delimiter=batch.delimiter or ',')
        
        # Skip header row
        next(csv_reader, None)
        
        # Process in chunks
        chunk_data = []
        row_number = 1
        
        for row in csv_reader:
            chunk_data.append(row)
            
            # Process chunk when it reaches the chunk size
            if len(chunk_data) >= chunk_size:
                chunk_success, chunk_errors = process_data_chunk(
                    batch, mapping, headers, chunk_data, row_number - len(chunk_data) + 1, task_instance,
                    existing_engine_keys, batch_engine_keys
                )
                
                success_count += chunk_success
                error_count += chunk_errors
                
                # Check if cancellation requested FIRST (before updating progress)
                # Use a fresh query to avoid any caching issues
                from django.db import connection
                with connection.cursor() as cursor:
                    cursor.execute(
                        "SELECT cancel_requested FROM imports_importbatch WHERE id = %s",
                        [batch.id]
                    )
                    row = cursor.fetchone()
                    if row and row[0]:  # cancel_requested is True
                        ImportLog.objects.create(
                            batch=batch,
                            level='warning',
                            message=f"Import cancelled by user at row {row_number}"
                        )
                        batch.status = 'cancelled'
                        batch.progress_percentage = min(100, int(row_number / batch.total_rows * 100)) if batch.total_rows > 0 else 0
                        batch.processed_rows = row_number
                        batch.cancel_requested = True
                        batch.save()
                        return success_count, error_count
                
                # Update progress
                progress = min(100, int(row_number / batch.total_rows * 100)) if batch.total_rows > 0 else 0
                batch.progress_percentage = progress
                batch.processed_rows = row_number
                batch.save()
                
                # Update Celery task state
                if task_instance and CELERY_AVAILABLE and hasattr(task_instance, 'update_state'):
                    task_instance.update_state(
                        state='PROGRESS',
                        meta={
                            'rows': row_number,
                            'created': success_count,
                            'errors': error_count,
                            'status': f'Processing row {row_number}...'
                        }
                    )
                
                # Clear chunk
                chunk_data = []
            
            row_number += 1
        
        # Process remaining data in chunk
        if chunk_data:
            chunk_success, chunk_errors = process_data_chunk(
                batch, mapping, headers, chunk_data, row_number - len(chunk_data), task_instance
            )
            success_count += chunk_success
            error_count += chunk_errors
    
    return success_count, error_count

def process_xlsx_import(batch, mapping, task_instance=None):
    """Process XLSX import in chunks with streaming."""
    success_count = 0
    error_count = 0
    
    # Preload existing engine keys for fast duplicate detection
    existing_engine_keys = set()
    for make, model, ident in Engine.objects.values_list('engine_make', 'engine_model', 'identifier'):
        existing_engine_keys.add((_nz(make), _nz(model), _nz(ident)))
    
    # Track keys we create/update in this run to avoid dupes within the batch
    batch_engine_keys = set()
    
    # Use streaming approach for XLSX
    chunk_size = mapping.chunk_size or 1000  # Default to 1000 if not set
    
    # Stream process XLSX file
    with batch.file.open('rb') as f:
        from openpyxl import load_workbook
        
        # Load workbook with read_only=True for memory efficiency
        workbook = load_workbook(f, read_only=True, data_only=True)
        
        # Get the specified worksheet
        worksheet_name = batch.worksheet_name or workbook.sheetnames[0]
        worksheet = workbook[worksheet_name]
        
        # Get headers from first row
        headers = []
        first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row:
            headers = [str(cell).strip() if cell is not None else '' for cell in first_row]
        
        # Process rows in chunks
        chunk_data = []
        row_number = 1
        
        # Iterate through rows starting from row 2 (skip header)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            # Convert row to list and handle None values
            row_data = [str(cell).strip() if cell is not None else '' for cell in row]
            
            # Skip completely empty rows
            if not any(cell for cell in row_data):
                continue
            
            chunk_data.append(row_data)
            
            # Process chunk when it reaches the chunk size
            if len(chunk_data) >= chunk_size:
                chunk_success, chunk_errors = process_data_chunk(
                    batch, mapping, headers, chunk_data, row_number - len(chunk_data) + 1, task_instance
                )
                
                success_count += chunk_success
                error_count += chunk_errors
                
                # Check if cancellation requested FIRST (before updating progress)
                # Use a fresh query to avoid any caching issues
                from django.db import connection
                with connection.cursor() as cursor:
                    cursor.execute(
                        "SELECT cancel_requested FROM imports_importbatch WHERE id = %s",
                        [batch.id]
                    )
                    row_result = cursor.fetchone()
                    if row_result and row_result[0]:  # cancel_requested is True
                        ImportLog.objects.create(
                            batch=batch,
                            level='warning',
                            message=f"Import cancelled by user at row {row_number}"
                        )
                        batch.status = 'cancelled'
                        batch.progress_percentage = min(100, int(row_number / batch.total_rows * 100)) if batch.total_rows > 0 else 0
                        batch.processed_rows = row_number
                        batch.cancel_requested = True
                        batch.save()
                        workbook.close()
                        return success_count, error_count
                
                # Update progress
                progress = min(100, int(row_number / batch.total_rows * 100)) if batch.total_rows > 0 else 0
                batch.progress_percentage = progress
                batch.processed_rows = row_number
                batch.save()
                
                # Update Celery task state
                if task_instance and CELERY_AVAILABLE and hasattr(task_instance, 'update_state'):
                    task_instance.update_state(
                        state='PROGRESS',
                        meta={
                            'rows': row_number,
                            'created': success_count,
                            'errors': error_count,
                            'status': f'Processing row {row_number}...'
                        }
                    )
                
                # Clear chunk
                chunk_data = []
            
            row_number += 1
        
        # Process remaining data in chunk
        if chunk_data:
            chunk_success, chunk_errors = process_data_chunk(
                batch, mapping, headers, chunk_data, row_number - len(chunk_data), task_instance
            )
            success_count += chunk_success
            error_count += chunk_errors
        
        workbook.close()
    
    return success_count, error_count

def process_data_chunk(batch, mapping, headers, chunk_data, start_row, task_instance=None, existing_engine_keys=None, batch_engine_keys=None):
    """Process a chunk of data rows with transaction.atomic() per chunk and bulk operations."""
    success_count = 0
    error_count = 0
    
    # Pre-create ImportRow records in bulk to reduce DB queries
    import_rows_to_create = []
    for row_index, row in enumerate(chunk_data):
        row_number = start_row + row_index
        row_data = dict(zip(headers, row))
        import_rows_to_create.append(ImportRow(
            batch=batch,
            row_number=row_number,
            original_data=row_data
        ))
    
    # Bulk create ImportRow records
    created_import_rows = ImportRow.objects.bulk_create(import_rows_to_create)
    
    # Process each row with chunked transactions
    for row_index, (row, import_row) in enumerate(zip(chunk_data, created_import_rows)):
        row_number = start_row + row_index
        row_start_time = time.time()
        row_data = dict(zip(headers, row))
        
        try:
            # Process row within transaction
            with transaction.atomic():
                # Normalize data
                normalized_data = normalize_row_data(row_data, mapping)
                
                # Update ImportRow with normalized data
                import_row.normalized_machine_data = normalized_data.get('machine', {})
                import_row.normalized_engine_data = normalized_data.get('engine', {})
                import_row.normalized_part_data = normalized_data.get('part', {})
                import_row.normalized_vendor_data = normalized_data.get('vendor', {})
                import_row.normalized_buildlist_data = normalized_data.get('buildlist', {})
                import_row.normalized_buildlistitem_data = normalized_data.get('buildlistitem', {})
                import_row.normalized_kit_data = normalized_data.get('kit', {})
                import_row.normalized_kititem_data = normalized_data.get('kititem', {})
                
                # Process each section based on mapping
                if mapping.machine_mapping and mapping.machine_mapping != {}:
                    process_machine_row(batch, mapping, normalized_data, import_row)
                
                if mapping.engine_mapping and mapping.engine_mapping != {}:
                    process_engine_row(batch, mapping, normalized_data, import_row, existing_engine_keys, batch_engine_keys)
                
                if mapping.part_mapping and mapping.part_mapping != {}:
                    process_part_row(batch, mapping, normalized_data, import_row)
                
                # Vendor processing is now handled within engine/part processing
                
                # Process build lists and kits
                if mapping.buildlist_mapping and mapping.buildlist_mapping != {}:
                    process_buildlist_row(batch, mapping, normalized_data, import_row)
                
                if mapping.kit_mapping and mapping.kit_mapping != {}:
                    process_kit_row(batch, mapping, normalized_data, import_row)
                
                # Create relationships
                create_relationships(batch, mapping, normalized_data, import_row)
                
                # Update processing time
                import_row.processing_time_ms = int((time.time() - row_start_time) * 1000)
                import_row.save()
                
                success_count += 1
                
        except Exception as e:
            error_count += 1
            # Handle error outside of transaction
            import_row.add_error(str(e))
            import_row.processing_time_ms = int((time.time() - row_start_time) * 1000)
            import_row.save()
            
            # Log error outside of transaction
            ImportLog.objects.create(
                batch=batch,
                level='error',
                message=f"Row {row_number}: {str(e)}",
                row_number=row_number
            )
    
    return success_count, error_count

def normalize_row_data(row_data: Dict[str, Any], mapping: SavedImportMapping) -> Dict[str, Dict[str, Any]]:
    """Normalize row data according to the specification."""
    normalized = {
        'machine': {},
        'engine': {},
        'part': {},
        'vendor': {},
        'buildlist': {},
        'buildlistitem': {},
        'kit': {},
        'kititem': {},
        'raw_data': row_data  # Include raw data for attribute processing
    }
    
    # Helper function to normalize a value
    def normalize_value(value, field_type='string'):
        if value is None or value == '':
            return None
        
        # Convert to string first
        value = str(value).strip()
        
        if value == '':
            return None
        
        # Collapse multiple spaces
        value = re.sub(r'\s+', ' ', value)
        
        if field_type == 'string':
            return value
        elif field_type == 'uppercase':
            return value.upper()
        elif field_type == 'integer':
            try:
                return int(float(value))  # Handle Excel numbers stored as text
            except (ValueError, TypeError):
                return None
        elif field_type == 'decimal':
            try:
                return Decimal(str(value))
            except (InvalidOperation, ValueError):
                return None
        elif field_type == 'boolean':
            # Handle various boolean formats
            value_lower = value.lower().strip()
            
            # Truthy values
            if value_lower in ('true', '1', 'yes', 'y', '✓'):
                return True
            # Falsy values
            elif value_lower in ('false', '0', 'no', 'n', 'x', ''):
                return False
            # For any other value, log warning and default to False
            else:
                # Log warning for invalid boolean values
                print(f"Warning: Invalid boolean value '{value}' - defaulting to False")
                return False
        
        return value
    
    # Normalize machine data
    for field_name, header_name in mapping.machine_mapping.items():
        if header_name in row_data:
            value = row_data[header_name]
            
            # Determine field type for normalization
            if field_name in ['make', 'model', 'machine_type', 'market_type']:
                normalized_value = normalize_value(value, 'string')
            elif field_name == 'year':
                normalized_value = normalize_value(value, 'integer')
            else:
                normalized_value = normalize_value(value, 'string')
            
            if normalized_value is not None:
                normalized['machine'][field_name] = normalized_value
    
    # Normalize engine data
    for field_name, header_name in mapping.engine_mapping.items():
        if header_name in row_data:
            value = row_data[header_name]
            
            # Determine field type for normalization
            if field_name in ['engine_make', 'engine_model', 'identifier', 'sg_engine_identifier', 'sg_engine_notes',
                             'cpl_number', 'ar_number', 'build_list', 'engine_code', 'serial_number',
                             'crankshaft_no', 'piston_no', 'piston_marked_no', 'piston_notes', 'oh_kit_no', 
                             'bore_stroke', 'firing_order', 'overview_comments', 'interference', 'camshaft', 
                             'valve_adjustment', 'status']:
                normalized_value = normalize_value(value, 'string')
            elif field_name in ['cylinder', 'valves_per_cyl']:
                normalized_value = normalize_value(value, 'integer')
            elif field_name in ['compression_ratio', 'rod_journal_diameter', 'main_journal_diameter_pos1',
                               'main_journal_diameter_1', 'big_end_housing_bore', 'price']:
                normalized_value = normalize_value(value, 'decimal')
            elif field_name in ['injection_type', 'valve_config', 'fuel_system_type']:
                normalized_value = normalize_value(value, 'string')
            else:
                normalized_value = normalize_value(value, 'string')
            
            if normalized_value is not None:
                normalized['engine'][field_name] = normalized_value
    
    # Normalize part data
    for field_name, header_name in mapping.part_mapping.items():
        if header_name in row_data:
            value = row_data[header_name]
            
            # Determine field type for normalization
            if field_name == 'part_number':
                normalized_value = normalize_value(value, 'uppercase')
            elif field_name in ['name', 'category', 'manufacturer', 'unit', 'type', 'manufacturer_type']:
                normalized_value = normalize_value(value, 'string')
            elif field_name == 'weight':
                normalized_value = normalize_value(value, 'decimal')
            else:
                normalized_value = normalize_value(value, 'string')
            
            if normalized_value is not None:
                normalized['part'][field_name] = normalized_value
    
    # Normalize vendor data
    for field_name, header_name in mapping.vendor_mapping.items():
        if header_name in row_data:
            value = row_data[header_name]
            
            # Determine field type for normalization
            if field_name in ['vendor_name', 'vendor_website', 'vendor_contact_name', 'vendor_contact_email', 'vendor_contact_phone', 'vendor_notes', 'vendor_part_number']:
                normalized_value = normalize_value(value, 'string')
            elif field_name in ['vendor_price']:
                normalized_value = normalize_value(value, 'decimal')
            elif field_name in ['vendor_stock_qty']:
                normalized_value = normalize_value(value, 'integer')
            else:
                normalized_value = normalize_value(value, 'string')
            
            if normalized_value is not None:
                normalized['vendor'][field_name] = normalized_value
    
    # Normalize build list data
    for field_name, header_name in mapping.buildlist_mapping.items():
        if header_name in row_data:
            value = row_data[header_name]
            
            # All build list fields are strings
            normalized_value = normalize_value(value, 'string')
            
            if normalized_value is not None:
                normalized['buildlist'][field_name] = normalized_value
    
    # Normalize build list item data
    for field_name, header_name in mapping.buildlistitem_mapping.items():
        if header_name in row_data:
            value = row_data[header_name]
            
            # Determine field type for normalization
            if field_name in ['name', 'description']:
                normalized_value = normalize_value(value, 'string')
            elif field_name == 'hour_qty':
                normalized_value = normalize_value(value, 'decimal')
            else:
                normalized_value = normalize_value(value, 'string')
            
            if normalized_value is not None:
                normalized['buildlistitem'][field_name] = normalized_value
    
    # Normalize kit data
    for field_name, header_name in mapping.kit_mapping.items():
        if header_name in row_data:
            value = row_data[header_name]
            
            # All kit fields are strings
            normalized_value = normalize_value(value, 'string')
            
            if normalized_value is not None:
                normalized['kit'][field_name] = normalized_value
    
    # Normalize kit item data
    for field_name, header_name in mapping.kititem_mapping.items():
        if header_name in row_data:
            value = row_data[header_name]
            
            # Determine field type for normalization
            if field_name == 'part_number':
                normalized_value = normalize_value(value, 'uppercase')
            elif field_name in ['quantity', 'part_weight']:
                normalized_value = normalize_value(value, 'decimal')
            elif field_name in ['part_name', 'part_category', 'part_manufacturer', 'part_unit', 'part_type', 'part_manufacturer_type']:
                normalized_value = normalize_value(value, 'string')
            else:
                normalized_value = normalize_value(value, 'string')
            
            if normalized_value is not None:
                normalized['kititem'][field_name] = normalized_value
    
    return normalized

def process_machine_row(batch, mapping, normalized_data, import_row):
    """Process a machine row with deduplication rules."""
    machine_data = normalized_data.get('machine', {})
    
    if not machine_data:
        return
    
    # Validate at least one field has a value
    if not any(v is not None and v != '' for v in machine_data.values()):
        error_msg = "Machine row must have at least one field with data"
        import_row.add_error(error_msg)
        raise Exception(error_msg)
    
    # Apply deduplication rules: (make, model, year, machine_type, market_type) CI
    # Only include non-null fields in the filter
    dedupe_fields = ['make', 'model', 'year', 'machine_type', 'market_type']
    dedupe_filters = {}
    
    # String fields that should use case-insensitive matching
    string_fields = ['make', 'model', 'machine_type', 'market_type']
    
    for field in dedupe_fields:
        if field in machine_data and machine_data[field] is not None:
            if field in string_fields:
                dedupe_filters[f"{field}__iexact"] = machine_data[field]
            else:
                # For non-string fields (like year), use exact matching
                dedupe_filters[field] = machine_data[field]
    
    # Check for existing machine - only if we have identifying information
    existing_machine = None
    if dedupe_filters:
        existing_machine = Machine.objects.filter(**dedupe_filters).first()
    
    if existing_machine:
        if mapping.skip_duplicates:
            # Skip duplicate
            import_row.machine_id = existing_machine.id
            import_row.save()
            return
        elif mapping.update_existing:
            # Update existing
            filtered_machine_data = filter_valid_fields(machine_data, Machine)
            validated_machine_data = validate_and_truncate_fields(filtered_machine_data, Machine, batch, import_row.row_number)
            for field, value in validated_machine_data.items():
                setattr(existing_machine, field, value)
            existing_machine.save()
            import_row.machine_updated = True
            import_row.machine_id = existing_machine.id
            import_row.save()
            return
    
    # Filter out fields that don't exist on the Machine model
    filtered_machine_data = filter_valid_fields(machine_data, Machine)
    
    # Validate and truncate field lengths
    validated_machine_data = validate_and_truncate_fields(filtered_machine_data, Machine, batch, import_row.row_number)
    
    # Create new machine
    machine = Machine.objects.create(**validated_machine_data)
    import_row.machine_created = True
    import_row.machine_id = machine.id
    import_row.save()

def process_engine_row(batch, mapping, normalized_data, import_row, existing_engine_keys=None, batch_engine_keys=None):
    """Process an engine row with deduplication rules based on (make, model, identifier) triple."""
    engine_data = normalized_data.get('engine', {})
    
    if not engine_data:
        return
    
    # Validate at least one field has a value
    if not any(v is not None and v != '' for v in engine_data.values()):
        error_msg = "Engine row must have at least one field with data"
        import_row.add_error(error_msg)
        raise Exception(error_msg)
    
    # Handle SG Engine mapping
    sg_engine = None
    if engine_data.get('sg_make') and engine_data.get('sg_model'):
        # Validate and truncate SGEngine data
        sg_engine_defaults = {
            'sg_make': engine_data['sg_make'],
            'sg_model': engine_data['sg_model'],
            'identifier': f"{engine_data['sg_make']}_{engine_data['sg_model']}".replace(' ', '_').upper()
        }
        validated_sg_defaults = validate_and_truncate_fields(sg_engine_defaults, SGEngine, batch, import_row.row_number)
        
        sg_engine, created = SGEngine.objects.get_or_create(
            sg_make__iexact=engine_data['sg_make'],
            sg_model__iexact=engine_data['sg_model'],
            defaults=validated_sg_defaults
        )
    else:
        # If no sg_make/sg_model provided, use Unknown default
        sg_engine = SGEngine.objects.get_or_create(
            identifier='UNKNOWN_DEFAULT',
            defaults={
                'sg_make': 'Unknown',
                'sg_model': 'Unknown',
                'notes': 'Default record for imports without engine information'
            }
        )[0]
    
    # Store sg_engine instance for later use (don't add to engine_data as it's not JSON serializable)
    # We'll add it to the filtered/validated data before creating the Engine object
    
    # Build engine key from row data using mapping
    row_data = {}
    for field_name, header_name in mapping.engine_mapping.items():
        if header_name in normalized_data.get('engine', {}):
            row_data[header_name] = normalized_data['engine'][header_name]
    
    eng_key = build_engine_key(row_data, mapping.engine_mapping)
    
    # Apply "Skip Duplicates" logic based on (make, model, identifier) triple
    skip_dupes = mapping.skip_duplicates
    update_existing = mapping.update_existing
    
    if skip_dupes and is_valid_engine_key(eng_key):
        if eng_key in (batch_engine_keys or set()) or eng_key in (existing_engine_keys or set()):
            if update_existing:
                # Fetch target and update instead of skipping
                target = Engine.objects.filter(
                    engine_make__iexact=eng_key[0],
                    engine_model__iexact=eng_key[1],
                    identifier__iexact=eng_key[2],
                ).first()
                if target:
                    # Update existing engine
                    filtered_engine_data = filter_valid_fields(engine_data, Engine)
                    validated_engine_data = validate_and_truncate_fields(filtered_engine_data, Engine, batch, import_row.row_number)
                    
                    # Add sg_engine to the validated data if we have one
                    if sg_engine:
                        validated_engine_data['sg_engine'] = sg_engine
                    
                    for field, value in validated_engine_data.items():
                        setattr(target, field, value)
                    target.save()
                    import_row.engine_updated = True
                    import_row.engine_duplicate_updated = True
                    import_row.engine_id = target.id
                    import_row.save()
                # Do not create a new Engine
                return
            else:
                # Skip duplicate
                import_row.engine_id = None  # Mark as skipped
                import_row.engine_duplicate_skipped = True
                import_row.save()
                return
    
    # Not a duplicate (or skip_dupes off) → create new engine
    # Filter out fields that don't exist on the Engine model
    filtered_engine_data = filter_valid_fields(engine_data, Engine)
    
    # Validate and truncate field lengths
    validated_engine_data = validate_and_truncate_fields(filtered_engine_data, Engine, batch, import_row.row_number)
    
    # Add sg_engine to the validated data (it's a ForeignKey, not a regular field)
    if sg_engine:
        validated_engine_data['sg_engine'] = sg_engine
    
    # Create new engine
    engine = Engine.objects.create(**validated_engine_data)
    
    # Attach vendor if present in row data
    vendor_data = normalized_data.get('vendor', {})
    if vendor_data.get('vendor_name'):
        vendor = ci_get_or_create_vendor(vendor_data['vendor_name'], vendor_data)
        if vendor:
            engine.vendor = vendor
            engine.save(update_fields=['vendor'])
    
    # After save, remember key so later rows see it as duplicate
    if is_valid_engine_key(eng_key):
        if batch_engine_keys is not None:
            batch_engine_keys.add(eng_key)
        if existing_engine_keys is not None:
            existing_engine_keys.add(eng_key)
    
    import_row.engine_created = True
    import_row.engine_id = engine.id
    import_row.save()

def process_part_row(batch, mapping, normalized_data, import_row):
    """Process a part row with deduplication rules."""
    part_data = normalized_data.get('part', {})
    
    if not part_data:
        return
    
    # Validate at least one field has a value
    if not any(v is not None and v != '' for v in part_data.values()):
        error_msg = "Part row must have at least one field with data"
        import_row.add_error(error_msg)
        raise Exception(error_msg)
    
    # Apply deduplication rules: (part_number, name) CI
    # Only include non-null fields in the filter
    dedupe_filters = {}
    if part_data.get('part_number') is not None:
        dedupe_filters['part_number__iexact'] = part_data['part_number']
    if part_data.get('name') is not None:
        dedupe_filters['name__iexact'] = part_data['name']
    
    # Check for existing part - only if we have identifying information
    existing_part = None
    if dedupe_filters:
        existing_part = Part.objects.filter(**dedupe_filters).first()
    
    if existing_part:
        if mapping.skip_duplicates:
            # Skip duplicate
            import_row.part_id = existing_part.id
            import_row.save()
            # Apply attributes to existing part
            apply_part_attributes_from_row(existing_part, normalized_data.get('raw_data', {}), mapping, batch, import_row)
            return
        elif mapping.update_existing:
            # Update existing
            filtered_part_data = filter_valid_fields(part_data, Part)
            validated_part_data = validate_and_truncate_fields(filtered_part_data, Part, batch, import_row.row_number)
            for field, value in validated_part_data.items():
                setattr(existing_part, field, value)
            existing_part.save()
            
            # Update vendor pricing if present
            vendor_data = normalized_data.get('vendor', {})
            if vendor_data.get('vendor_name'):
                vendor = ci_get_or_create_vendor(vendor_data['vendor_name'], vendor_data)
                if vendor:
                    existing_part.vendor = vendor
                    existing_part.save(update_fields=['vendor'])
                    
                    # Create or update PartVendor relationship with pricing
                    part_vendor, created = PartVendor.objects.get_or_create(
                        part=existing_part, 
                        vendor=vendor,
                        defaults={
                            'vendor_part_number': vendor_data.get('vendor_part_number', ''),
                            'price': vendor_data.get('vendor_price'),
                            'stock_qty': vendor_data.get('vendor_stock_qty'),
                        }
                    )
                    
                    # Update existing PartVendor if not created
                    if not created:
                        if vendor_data.get('vendor_part_number'):
                            part_vendor.vendor_part_number = vendor_data['vendor_part_number']
                        if vendor_data.get('vendor_price') is not None:
                            part_vendor.price = vendor_data['vendor_price']
                        if vendor_data.get('vendor_stock_qty') is not None:
                            part_vendor.stock_qty = vendor_data['vendor_stock_qty']
                        part_vendor.save()
                    
                    # Track part vendor relationship creation
                    if created:
                        import_row.part_vendor_created = True
            
            import_row.part_updated = True
            import_row.part_id = existing_part.id
            import_row.save()
            # Apply attributes to existing part
            apply_part_attributes_from_row(existing_part, normalized_data.get('raw_data', {}), mapping, batch, import_row)
            return
    
    # Filter out fields that don't exist on the Part model
    filtered_part_data = filter_valid_fields(part_data, Part)
    
    # Validate and truncate field lengths
    validated_part_data = validate_and_truncate_fields(filtered_part_data, Part, batch, import_row.row_number)
    
    # Create new part
    part = Part.objects.create(**validated_part_data)
    
    # Attach vendor if present in row data
    vendor_data = normalized_data.get('vendor', {})
    if vendor_data.get('vendor_name'):
        vendor = ci_get_or_create_vendor(vendor_data['vendor_name'], vendor_data)
        if vendor:
            part.vendor = vendor
            part.save(update_fields=['vendor'])
            
            # Create or update PartVendor relationship with pricing
            part_vendor, created = PartVendor.objects.get_or_create(
                part=part, 
                vendor=vendor,
                defaults={
                    'vendor_part_number': vendor_data.get('vendor_part_number', ''),
                    'price': vendor_data.get('vendor_price'),
                    'stock_qty': vendor_data.get('vendor_stock_qty'),
                }
            )
            
            # Update existing PartVendor if not created
            if not created:
                if vendor_data.get('vendor_part_number'):
                    part_vendor.vendor_part_number = vendor_data['vendor_part_number']
                if vendor_data.get('vendor_price') is not None:
                    part_vendor.price = vendor_data['vendor_price']
                if vendor_data.get('vendor_stock_qty') is not None:
                    part_vendor.stock_qty = vendor_data['vendor_stock_qty']
                part_vendor.save()
            
            # Track part vendor relationship creation
            if created:
                import_row.part_vendor_created = True
            
            # Auto-set primary vendor if only one vendor exists
            part.auto_set_primary_vendor()
    
    import_row.part_created = True
    import_row.part_id = part.id
    import_row.save()
    
    # Apply attributes to new part
    apply_part_attributes_from_row(part, normalized_data.get('raw_data', {}), mapping, batch, import_row)

def apply_part_attributes_from_row(part, row_dict, mapping, batch, import_row):
    """Apply mapped part attributes from CSV row data."""
    attr_map = mapping.part_attribute_mappings or {}
    if not attr_map:
        return
    
    # Build a set of allowed attribute IDs for the part's category
    allowed_ids = set(
        PartAttribute.objects
        .filter(category=part.category)
        .values_list("id", flat=True)
    )
    
    for attr_id_str, csv_header in attr_map.items():
        try:
            attr_id = int(attr_id_str)
        except ValueError:
            continue
        
        raw_val = row_dict.get(csv_header, "")
        if raw_val in (None, ""):
            continue
        
        if attr_id not in allowed_ids:
            # Log skip
            ImportLog.objects.create(
                batch=batch,
                level='warning',
                message=f"Skipped attribute {attr_id} — not in category {part.category}",
                row_number=import_row.row_number
            )
            continue
        
        try:
            attr = PartAttribute.objects.get(pk=attr_id)
            
            # Get or create PartAttributeValue
            pav, created = PartAttributeValue.objects.get_or_create(
                part=part, 
                attribute=attr,
                defaults={}
            )
            
            # Clear all typed fields
            pav.value_text = None
            pav.value_int = None
            pav.value_dec = None
            pav.value_bool = None
            pav.value_date = None
            pav.choice = None
            
            # Set the appropriate typed field based on attribute.data_type
            if attr.data_type == "text":
                pav.value_text = str(raw_val)
            elif attr.data_type == "int":
                try:
                    pav.value_int = int(str(raw_val).strip())
                except ValueError:
                    ImportLog.objects.create(
                        batch=batch,
                        level='warning',
                        message=f"Invalid integer value '{raw_val}' for attribute {attr.name}",
                        row_number=import_row.row_number
                    )
                    continue
            elif attr.data_type == "dec":
                try:
                    from decimal import Decimal
                    pav.value_dec = Decimal(str(raw_val))
                except (ValueError, InvalidOperation):
                    ImportLog.objects.create(
                        batch=batch,
                        level='warning',
                        message=f"Invalid decimal value '{raw_val}' for attribute {attr.name}",
                        row_number=import_row.row_number
                    )
                    continue
            elif attr.data_type == "bool":
                pav.value_bool = str(raw_val).strip().lower() in {"1", "true", "yes", "y"}
            elif attr.data_type == "date":
                try:
                    from datetime import datetime
                    # Try common date formats
                    for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S']:
                        try:
                            pav.value_date = datetime.strptime(str(raw_val).strip(), fmt).date()
                            break
                        except ValueError:
                            continue
                    else:
                        ImportLog.objects.create(
                            batch=batch,
                            level='warning',
                            message=f"Invalid date format '{raw_val}' for attribute {attr.name}",
                            row_number=import_row.row_number
                        )
                        continue
                except Exception:
                    ImportLog.objects.create(
                        batch=batch,
                        level='warning',
                        message=f"Invalid date value '{raw_val}' for attribute {attr.name}",
                        row_number=import_row.row_number
                    )
                    continue
            elif attr.data_type == "choice":
                choice = attr.choices.filter(value=str(raw_val)).first()
                if choice:
                    pav.choice = choice
                else:
                    ImportLog.objects.create(
                        batch=batch,
                        level='warning',
                        message=f"Invalid choice value '{raw_val}' for attribute {attr.name}",
                        row_number=import_row.row_number
                    )
                    continue
            
            pav.save()
            
        except PartAttribute.DoesNotExist:
            ImportLog.objects.create(
                batch=batch,
                level='warning',
                message=f"Attribute {attr_id} not found",
                row_number=import_row.row_number
            )
            continue
        except Exception as e:
            ImportLog.objects.create(
                batch=batch,
                level='error',
                message=f"Error setting attribute {attr_id}: {str(e)}",
                row_number=import_row.row_number
            )
            continue


def process_buildlist_row(batch, mapping, normalized_data, import_row):
    """Process a build list row with deduplication rules."""
    from inventory.models import BuildList, BuildListItem
    
    buildlist_data = normalized_data.get('buildlist', {})
    buildlistitem_data = normalized_data.get('buildlistitem', {})
    
    # Must have at least build list name or item data
    if not buildlist_data and not buildlistitem_data:
        return
    
    # Get or create build list by name (case-insensitive)
    buildlist_name = buildlist_data.get('name')
    if not buildlist_name:
        error_msg = "Build list name is required"
        import_row.add_error(error_msg)
        raise Exception(error_msg)
    
    # Check for existing build list (case-insensitive)
    existing_buildlist = BuildList.objects.filter(name__iexact=buildlist_name).first()
    
    if existing_buildlist:
        if mapping.skip_duplicates and not mapping.update_existing:
            # Skip - just track it
            import_row.buildlist_id = existing_buildlist.id
            buildlist = existing_buildlist
        elif mapping.update_existing:
            # Update existing
            if buildlist_data.get('notes'):
                existing_buildlist.notes = buildlist_data['notes']
            existing_buildlist.save()
            import_row.buildlist_updated = True
            import_row.buildlist_id = existing_buildlist.id
            buildlist = existing_buildlist
        else:
            # Not skipping, create with duplicate name
            buildlist = BuildList.objects.create(
                name=buildlist_name,
                notes=buildlist_data.get('notes', ''),
                created_by=batch.created_by
            )
            import_row.buildlist_created = True
            import_row.buildlist_id = buildlist.id
    else:
        # Create new build list
        buildlist = BuildList.objects.create(
            name=buildlist_name,
            notes=buildlist_data.get('notes', ''),
            created_by=batch.created_by
        )
        import_row.buildlist_created = True
        import_row.buildlist_id = buildlist.id
    
    # Create build list item if item data present
    if buildlistitem_data and buildlistitem_data.get('name'):
        item_name = buildlistitem_data.get('name')
        item_description = buildlistitem_data.get('description', '')
        item_hour_qty = buildlistitem_data.get('hour_qty', Decimal('0.00'))
        
        # Create build list item
        buildlistitem = BuildListItem.objects.create(
            build_list=buildlist,
            name=item_name,
            description=item_description,
            hour_qty=item_hour_qty
        )
        import_row.buildlistitem_created = True
        import_row.buildlistitem_id = buildlistitem.id
    
    import_row.save()


def process_kit_row(batch, mapping, normalized_data, import_row):
    """Process a kit row with deduplication rules and part auto-creation."""
    from inventory.models import Kit, KitItem, Part, PartCategory
    
    kit_data = normalized_data.get('kit', {})
    kititem_data = normalized_data.get('kititem', {})
    
    # Must have at least kit name or item data
    if not kit_data and not kititem_data:
        return
    
    # Get or create kit by name (case-insensitive)
    kit_name = kit_data.get('name')
    if not kit_name:
        error_msg = "Kit name is required"
        import_row.add_error(error_msg)
        raise Exception(error_msg)
    
    # Check for existing kit (case-insensitive)
    existing_kit = Kit.objects.filter(name__iexact=kit_name).first()
    
    if existing_kit:
        if mapping.skip_duplicates and not mapping.update_existing:
            # Skip - just track it
            import_row.kit_id = existing_kit.id
            kit = existing_kit
        elif mapping.update_existing:
            # Update existing
            if kit_data.get('notes'):
                existing_kit.notes = kit_data['notes']
            existing_kit.save()
            import_row.kit_updated = True
            import_row.kit_id = existing_kit.id
            kit = existing_kit
        else:
            # Not skipping, create with duplicate name
            kit = Kit.objects.create(
                name=kit_name,
                notes=kit_data.get('notes', ''),
                created_by=batch.created_by
            )
            import_row.kit_created = True
            import_row.kit_id = kit.id
    else:
        # Create new kit
        kit = Kit.objects.create(
            name=kit_name,
            notes=kit_data.get('notes', ''),
            created_by=batch.created_by
        )
        import_row.kit_created = True
        import_row.kit_id = kit.id
    
    # Create kit item if item data present
    if kititem_data and kititem_data.get('part_number'):
        part_number = kititem_data.get('part_number')
        quantity = kititem_data.get('quantity', Decimal('1.00'))
        
        # Check if part exists
        part = Part.objects.filter(part_number__iexact=part_number).first()
        
        if not part:
            # Auto-create part with data from row
            part_data = {
                'part_number': part_number,
                'name': kititem_data.get('part_name', part_number),
            }
            
            # Add optional fields if present
            if kititem_data.get('part_manufacturer'):
                part_data['manufacturer'] = kititem_data['part_manufacturer']
            if kititem_data.get('part_unit'):
                part_data['unit'] = kititem_data['part_unit']
            if kititem_data.get('part_type'):
                part_data['type'] = kititem_data['part_type']
            if kititem_data.get('part_manufacturer_type'):
                part_data['manufacturer_type'] = kititem_data['part_manufacturer_type']
            if kititem_data.get('part_weight'):
                part_data['weight'] = kititem_data['part_weight']
            
            # Handle category (need to get or create PartCategory)
            if kititem_data.get('part_category'):
                category_name = kititem_data['part_category']
                # Try to find existing category (case-insensitive)
                category = PartCategory.objects.filter(name__iexact=category_name).first()
                if category:
                    part_data['category'] = category
            
            # Filter and validate part data
            filtered_part_data = filter_valid_fields(part_data, Part)
            validated_part_data = validate_and_truncate_fields(filtered_part_data, Part, batch, import_row.row_number)
            
            # Create new part
            part = Part.objects.create(**validated_part_data)
            import_row.part_auto_created = True
            import_row.part_created = True
            import_row.part_id = part.id
        
        # Create kit item linking kit to part
        # Check if kit item already exists for this kit and part
        existing_kititem = KitItem.objects.filter(kit=kit, part=part).first()
        if not existing_kititem:
            kititem = KitItem.objects.create(
                kit=kit,
                part=part,
                quantity=quantity
            )
            import_row.kititem_created = True
            import_row.kititem_id = kititem.id
        else:
            # Update quantity if update_existing is enabled
            if mapping.update_existing:
                existing_kititem.quantity = quantity
                existing_kititem.save()
            import_row.kititem_id = existing_kititem.id
    
    import_row.save()


def create_relationships(batch, mapping, normalized_data, import_row):
    """Create Machine↔Engine and Engine↔Part relationships."""
    
    # Create Machine↔Engine relationship
    if import_row.machine_id and import_row.engine_id:
        machine = Machine.objects.get(id=import_row.machine_id)
        engine = Engine.objects.get(id=import_row.engine_id)
        
        machine_engine, created = MachineEngine.objects.get_or_create(
            machine=machine,
            engine=engine,
            defaults={'notes': f"Created from import batch {batch.id}"}
        )
        
        if created:
            import_row.machine_engine_created = True
    
    # Create Engine↔Part relationship
    if import_row.engine_id and import_row.part_id:
        engine = Engine.objects.get(id=import_row.engine_id)
        part = Part.objects.get(id=import_row.part_id)
        
        engine_part, created = EnginePart.objects.get_or_create(
            engine=engine,
            part=part,
            defaults={'notes': f"Created from import batch {batch.id}"}
        )
        
        if created:
            import_row.engine_part_created = True
    
    # Engine↔Vendor relationship is now handled directly in process_engine_row
    
    # Create Machine↔Part relationship
    if import_row.machine_id and import_row.part_id:
        machine = Machine.objects.get(id=import_row.machine_id)
        part = Part.objects.get(id=import_row.part_id)
        
        machine_part, created = MachinePart.objects.get_or_create(
            machine=machine,
            part=part,
            defaults={'notes': f"Created from import batch {batch.id}"}
        )
        
        if created:
            import_row.machine_part_created = True
    
    # Handle vendor relationships
    vendor_data = normalized_data.get('vendor', {})
    if vendor_data and import_row.part_id:
        part = Part.objects.get(id=import_row.part_id)
        
        # Create or get vendor
        vendor_name = vendor_data.get('vendor_name')
        if vendor_name:
            # Validate vendor defaults
            vendor_defaults = {
                'name': vendor_name,
                'website': vendor_data.get('vendor_website', ''),
                'contact_name': vendor_data.get('vendor_contact_name', ''),
                'email': vendor_data.get('vendor_contact_email', ''),
                'phone': vendor_data.get('vendor_contact_phone', ''),
                'notes': vendor_data.get('vendor_notes', ''),
            }
            validated_vendor_defaults = validate_and_truncate_fields(vendor_defaults, Vendor, batch, import_row.row_number)
            
            vendor, created = Vendor.objects.get_or_create(
                name__iexact=vendor_name,
                defaults=validated_vendor_defaults
            )
            
            # Create PartVendor relationship
            part_vendor_data = {}
            if vendor_data.get('vendor_sku'):
                part_vendor_data['vendor_sku'] = vendor_data['vendor_sku']
            if vendor_data.get('vendor_cost'):
                part_vendor_data['cost'] = vendor_data['vendor_cost']
            if vendor_data.get('vendor_stock_qty'):
                part_vendor_data['stock_qty'] = vendor_data['vendor_stock_qty']
            if vendor_data.get('vendor_lead_time_days'):
                part_vendor_data['lead_time_days'] = vendor_data['vendor_lead_time_days']
            if vendor_data.get('vendor_notes'):
                part_vendor_data['notes'] = vendor_data['vendor_notes']
            
            # Validate PartVendor data
            validated_part_vendor_data = validate_and_truncate_fields(part_vendor_data, PartVendor, batch, import_row.row_number)
            
            part_vendor, created = PartVendor.objects.get_or_create(
                part=part,
                vendor=vendor,
                defaults=validated_part_vendor_data
            )
            
            if created:
                import_row.part_vendor_created = True
            
            # Set as primary vendor if specified
            if vendor_data.get('primary_vendor_name') == vendor_name:
                part.primary_vendor = vendor
                part.save()
            
            # Auto-set primary vendor if only one vendor exists
            part.auto_set_primary_vendor()
    
    import_row.save()
