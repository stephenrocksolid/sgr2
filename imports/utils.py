import csv
import io
import chardet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from django.core.files.base import ContentFile
from django.core.exceptions import ValidationError

def detect_encoding(file_content):
    """Detect the encoding of a file content."""
    result = chardet.detect(file_content)
    return result['encoding'] if result['confidence'] > 0.7 else 'utf-8'

def process_csv_file(file_content, encoding='utf-8', delimiter=','):
    """Process CSV file and return headers and preview data."""
    try:
        # Decode content
        if isinstance(file_content, bytes):
            content = file_content.decode(encoding, errors='replace')
        else:
            content = file_content
        
        # Create CSV reader
        csv_reader = csv.reader(io.StringIO(content), delimiter=delimiter)
        
        # Get headers (first row)
        headers = next(csv_reader, [])
        if not headers:
            raise ValidationError("CSV file appears to be empty or has no headers.")
        
        # Get preview data (next 200 rows)
        preview_data = []
        total_rows = 1  # Count header row
        
        for i, row in enumerate(csv_reader):
            if i >= 200:  # Limit preview to 200 rows
                break
            preview_data.append(row)
            total_rows += 1
        
        # Count remaining rows for total
        remaining_rows = sum(1 for _ in csv_reader)
        total_rows += remaining_rows
        
        return {
            'headers': headers,
            'preview_data': preview_data,
            'total_rows': total_rows
        }
    
    except UnicodeDecodeError:
        raise ValidationError(f"Unable to decode file with {encoding} encoding. Try a different encoding.")
    except Exception as e:
        raise ValidationError(f"Error processing CSV file: {str(e)}")

def process_csv_file_all_data(file_content, encoding='utf-8', delimiter=','):
    """Process CSV file and return headers and all data."""
    try:
        # Decode content
        if isinstance(file_content, bytes):
            content = file_content.decode(encoding, errors='replace')
        else:
            content = file_content
        
        # Create CSV reader
        csv_reader = csv.reader(io.StringIO(content), delimiter=delimiter)
        
        # Get headers (first row)
        headers = next(csv_reader, [])
        if not headers:
            raise ValidationError("CSV file appears to be empty or has no headers.")
        
        # Get all data rows
        data = []
        total_rows = 1  # Count header row
        
        for row in csv_reader:
            data.append(row)
            total_rows += 1
        
        return {
            'headers': headers,
            'data': data,
            'total_rows': total_rows
        }
    
    except UnicodeDecodeError:
        raise ValidationError(f"Unable to decode file with {encoding} encoding. Try a different encoding.")
    except Exception as e:
        raise ValidationError(f"Error processing CSV file: {str(e)}")

def process_xlsx_file(file_content):
    """Process XLSX file and return worksheet info and preview data."""
    try:
        # Load workbook
        workbook = load_workbook(
            filename=ContentFile(file_content),
            read_only=True,
            data_only=True
        )
        
        # Get worksheet names
        worksheet_names = workbook.sheetnames
        
        if not worksheet_names:
            raise ValidationError("Excel file contains no worksheets.")
        
        # Process first worksheet for preview
        worksheet = workbook[worksheet_names[0]]
        
        # Get headers (first row)
        headers = []
        for cell in worksheet[1]:
            headers.append(str(cell.value) if cell.value is not None else '')
        
        # Remove empty trailing headers
        while headers and not headers[-1].strip():
            headers.pop()
        
        if not headers:
            raise ValidationError("Excel file appears to be empty or has no headers.")
        
        # Get preview data (next 200 rows)
        preview_data = []
        total_rows = 1  # Count header row
        
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if len(preview_data) >= 200:  # Limit preview to 200 rows
                break
            preview_data.append(row)
            total_rows += 1
        
        # Count remaining rows for total
        for _ in worksheet.iter_rows(min_row=len(preview_data) + 2, values_only=True):
            total_rows += 1
        
        workbook.close()
        
        return {
            'headers': headers,
            'preview_data': preview_data,
            'total_rows': total_rows,
            'worksheet_names': worksheet_names
        }
    
    except Exception as e:
        raise ValidationError(f"Error processing Excel file: {str(e)}")

def get_xlsx_worksheet_data(file_content, worksheet_name):
    """Get data from a specific worksheet in an XLSX file."""
    try:
        workbook = load_workbook(
            filename=ContentFile(file_content),
            read_only=True,
            data_only=True
        )
        
        if worksheet_name not in workbook.sheetnames:
            raise ValidationError(f"Worksheet '{worksheet_name}' not found in file.")
        
        worksheet = workbook[worksheet_name]
        
        # Get headers
        headers = []
        for cell in worksheet[1]:
            headers.append(str(cell.value) if cell.value is not None else '')
        
        # Remove empty trailing headers
        while headers and not headers[-1].strip():
            headers.pop()
        
        # Get all data
        data = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            data.append(row)
        
        workbook.close()
        
        return {
            'headers': headers,
            'data': data,
            'total_rows': len(data) + 1  # +1 for header row
        }
    
    except Exception as e:
        raise ValidationError(f"Error reading worksheet '{worksheet_name}': {str(e)}")

def validate_file_limits(total_rows, file_size):
    """Validate file size and row count limits."""
    if file_size > 50 * 1024 * 1024:  # 50MB
        raise ValidationError("File size must be less than 50MB.")
    
    if total_rows > 100000:  # 100k rows
        raise ValidationError("File must have less than 100,000 rows.")
    
    if total_rows < 2:  # At least header + 1 data row
        raise ValidationError("File must contain at least one data row.")

def create_mapping_dict(form_data, section):
    """Create mapping dictionary from form data."""
    mapping = {}
    prefix = f'map_{section}_'
    
    for key, value in form_data.items():
        if key.startswith(prefix) and value:
            field_name = key[len(prefix):]
            mapping[field_name] = value
    
    return mapping

def get_expected_fields(section):
    """Get expected fields for a given section."""
    section_fields = {
        'machines': [
            ('make', 'Make'),
            ('model', 'Model'),
            ('year', 'Year'),
            ('machine_type', 'Machine Type'),
            ('market_type', 'Market Type'),
            ('serial_number', 'Serial Number'),
            ('notes', 'Notes'),
        ],
        'engines': [
            ('engine_make', 'Engine Make'),
            ('engine_model', 'Engine Model'),
            ('sg_make', 'SG Make'),
            ('sg_model', 'SG Model'),
            ('status', 'Status'),
            ('notes', 'Notes'),
        ],
        'parts': [
            ('part_number', 'Part Number'),
            ('name', 'Name'),
            ('manufacturer', 'Manufacturer'),
            ('category', 'Category'),
            ('description', 'Description'),
            ('notes', 'Notes'),
        ]
    }
    
    return section_fields.get(section, [])

def get_engine_field_aliases():
    """Get field aliases for engine import mapping."""
    return {
        'serial_number': ['s/n', 'serial', 'serial number', 'sn'],
        'di': ['di', 'direct injection'],
        'idi': ['idi', 'indirect injection'],
        'common_rail': ['common rail', 'common-rail', 'cr'],
        'two_valve': ['2v', '2 valve', 'two valve'],
        'four_valve': ['4v', '4 valve', 'four valve'],
        'five_valve': ['5v', '5 valve', 'five valve'],
        'casting_comments': ['casting # comments', 'casting comments', 'casting notes'],
    }

def fuzzy_match_header(header, target_field, aliases):
    """Fuzzy match a header to a target field using aliases."""
    header_lower = header.lower().strip()
    
    # Direct match
    if header_lower == target_field.lower():
        return True
    
    # Check aliases
    if target_field in aliases:
        for alias in aliases[target_field]:
            if header_lower == alias.lower():
                return True
            # Handle punctuation variations
            alias_clean = alias.lower().replace('#', '').replace('-', ' ').replace('_', ' ')
            header_clean = header_lower.replace('#', '').replace('-', ' ').replace('_', ' ')
            if alias_clean == header_clean:
                return True
    
    return False

def suggest_engine_field_mappings(headers):
    """Suggest field mappings for engine import based on header analysis."""
    aliases = get_engine_field_aliases()
    suggestions = {}
    
    for header in headers:
        for field, field_aliases in aliases.items():
            if fuzzy_match_header(header, field, aliases):
                suggestions[field] = header
                break
    
    return suggestions
